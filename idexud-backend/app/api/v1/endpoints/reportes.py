import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.db.session import get_db
from app.models.poliza import Poliza

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

# QUITAMOS EL PREFIJO AQUÍ PARA EVITAR EL 404
router = APIRouter(tags=["Reportes"])

COLOR_HEADER_BG   = "1A3A5C"
COLOR_HEADER_FG   = "FFFFFF"
COLOR_FILA_PAR    = "EEF4FB"
COLOR_FILA_IMPAR  = "FFFFFF"
COLOR_BORDE       = "B8C9D9"

CONTRASENA_PROTECCION = "idexud2024"

def _fmt_cop(valor: Optional[Decimal | float]) -> str:
    if valor is None: return "—"
    try: return f"{float(valor):,.0f}".replace(",", ".")
    except: return "—"

def _fmt_fecha(valor) -> str:
    if not valor: return "—"
    if isinstance(valor, (date, datetime)): return valor.strftime("%d/%m/%Y")
    return str(valor)

@router.get("/excel", summary="Exportar cartera a Excel")
async def exportar_excel(db: AsyncSession = Depends(get_db)):
    # 1. CONSULTA REAL A LA BASE DE DATOS
    result = await db.execute(
        select(Poliza)
        .options(selectinload(Poliza.aseguradora), selectinload(Poliza.contratista))
        .order_by(Poliza.created_at.desc())
    )
    polizas = result.scalars().all()

    # 2. CREAR EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera de Pólizas"

    # Estilos
    font_header = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=10)
    fill_header = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    fill_par = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    fill_impar = PatternFill("solid", fgColor=COLOR_FILA_IMPAR)
    lado = Side(style="thin", color=COLOR_BORDE)
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)

    # Cabeceras
    COLUMNAS = [
        ("N° Póliza / Radicado", 22), ("Estado", 16), ("Tipo", 20),
        ("Aseguradora", 25), ("Contratista", 25), ("Vigencia Hasta", 15),
        ("Valor Asegurado", 20), ("Valor Prima", 20)
    ]
    
    for col_idx, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.font, celda.fill, celda.border = font_header, fill_header, borde
        celda.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Filas de Datos Reales
    for fila_idx, p in enumerate(polizas, start=2):
        es_par = fila_idx % 2 == 0
        fill_fila = fill_par if es_par else fill_impar

        numero = p.numero_poliza or f"BORRADOR (ID: {p.id})"
        nom_aseg = p.aseguradora.nombre if p.aseguradora else "—"
        nom_cont = p.contratista.nombre_razon_social if p.contratista else "—"

        valores = [
            numero, 
            p.estado.value if p.estado else "—", 
            p.tipo.value if p.tipo else "—",
            nom_aseg, 
            nom_cont,
            _fmt_fecha(p.vigencia_hasta), 
            _fmt_cop(p.valor_asegurado), 
            _fmt_cop(p.valor_prima)
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.fill, celda.border = fill_fila, borde
            celda.font = Font(name="Arial", size=9, color="1A1A2E")
            celda.alignment = Alignment(horizontal="center" if col_idx in (2, 6) else "left")

    # 3. BLINDAJE (SOLO LECTURA)
    ws.protection = SheetProtection(password=CONTRASENA_PROTECCION, sheet=True)
    
    # 4. ENVIAR ARCHIVO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    nombre_archivo = f"cartera_polizas_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'}
    )